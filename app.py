import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import io
import os
import json
import re
import hashlib
import base64

# --- Configuración de persistencia de informes ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
REPORTS_DIR = os.path.join(BASE_DIR, "reports")
REPORTS_INDEX_FILE = os.path.join(REPORTS_DIR, "reports_index.json")
USERS_FILE = os.path.join(BASE_DIR, "users.json")

# --- Configuración de fechas ---
DATE_MIN_ALLOWED = datetime(2000, 1, 1)
DATE_MAX_FUTURE_YEARS = 3

DATE_FORMAT_OPTIONS = {
    "Auto (detectar)": "auto",
    "Excel serial (1900)": "excel_1900",
    "Excel serial (1904)": "excel_1904",
    "ISO (YYYY-MM-DD)": "iso",
    "DD/MM/YYYY": "dmy",
    "MM/DD/YYYY": "mdy",
    "YYYYMMDD": "yyyymmdd",
}

def load_reports_index():
    if not os.path.exists(REPORTS_DIR):
        os.makedirs(REPORTS_DIR)
    if os.path.exists(REPORTS_INDEX_FILE):
        with open(REPORTS_INDEX_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {} # Retorna un diccionario vacío si no existe el índice

def save_reports_index(index_data):
    if not os.path.exists(REPORTS_DIR):
        os.makedirs(REPORTS_DIR)
    with open(REPORTS_INDEX_FILE, "w", encoding="utf-8") as f:
        json.dump(index_data, f, indent=4)

def get_report_filepath(report_name_base):
    # Genera un nombre de archivo seguro y único para el Excel
    # Reemplazamos caracteres no válidos para nombres de archivo con guiones bajos
    safe_name = "".join(c if c.isalnum() else "_" for c in report_name_base).lower()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(REPORTS_DIR, f"{safe_name}_{timestamp}.xlsx")

def make_json_safe(obj):
    """Convierte tipos numpy/pandas a tipos JSON serializables."""
    if isinstance(obj, dict):
        return {k: make_json_safe(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [make_json_safe(v) for v in obj]
    if isinstance(obj, np.generic):
        return obj.item()
    if isinstance(obj, (pd.Timestamp, datetime)):
        return obj.isoformat()
    return obj

def safe_rerun():
    """Compatibilidad con versiones de Streamlit antiguas y nuevas."""
    if hasattr(st, "rerun"):
        st.rerun()
    elif hasattr(st, "experimental_rerun"):
        st.experimental_rerun()

def help_popup(title, render_fn):
    """Muestra ayuda en popup si está disponible; fallback a expander."""
    if hasattr(st, "popover"):
        with st.popover(title):
            render_fn()
    else:
        with st.expander(title, expanded=False):
            render_fn()

# --- Control de acceso ---
def hash_password(password: str, salt: str) -> str:
    return hashlib.sha256(f"{salt}{password}".encode("utf-8")).hexdigest()

def _normalize_users(users_raw):
    if users_raw is None:
        return {}
    if isinstance(users_raw, str):
        try:
            users_raw = json.loads(users_raw)
        except Exception:
            return {}
    if not isinstance(users_raw, dict):
        return {}
    # Permite formato {"users": {...}} o directamente {user: {salt, hash}}
    if "users" in users_raw and isinstance(users_raw["users"], dict):
        users_raw = users_raw["users"]
    users = {}
    for username, data in users_raw.items():
        if isinstance(data, dict) and "salt" in data and "hash" in data:
            users[username] = {"salt": str(data["salt"]), "hash": str(data["hash"])}
        elif isinstance(data, str):
            # Hash directo sin salt (no recomendado, pero soportado)
            users[username] = {"salt": "", "hash": data}
    return users

def load_users_config():
    users_raw = None
    env_b64 = os.getenv("CLV_USERS_B64")
    env_json = os.getenv("CLV_USERS_JSON")
    if env_b64:
        try:
            decoded = base64.b64decode(env_b64).decode("utf-8")
            users_raw = decoded
        except Exception:
            users_raw = None
    if users_raw is None and env_json:
        users_raw = env_json
    if users_raw is None:
        try:
            if "users" in st.secrets:
                users_raw = st.secrets["users"]
        except Exception:
            users_raw = None
    if users_raw is None and os.path.exists(USERS_FILE):
        with open(USERS_FILE, "r", encoding="utf-8") as f:
            try:
                users_raw = json.load(f)
            except json.JSONDecodeError:
                users_raw = None
    return _normalize_users(users_raw)

def verify_user(username: str, password: str, users: dict) -> bool:
    if username not in users:
        return False
    entry = users[username]
    salt = entry.get("salt", "")
    expected = entry.get("hash", "")
    return hash_password(password, salt) == expected

def require_auth():
    users = load_users_config()
    if not users:
        diagnostics = []
        env_b64 = os.getenv("CLV_USERS_B64")
        env_json = os.getenv("CLV_USERS_JSON")
        if env_b64:
            try:
                decoded = base64.b64decode(env_b64).decode("utf-8")
                json.loads(decoded)
                diagnostics.append("CLV_USERS_B64 presente y válido.")
            except Exception:
                diagnostics.append("CLV_USERS_B64 presente pero inválido (base64 o JSON).")
        if env_json:
            try:
                json.loads(env_json)
                diagnostics.append("CLV_USERS_JSON presente y válido.")
            except Exception:
                diagnostics.append("CLV_USERS_JSON presente pero inválido (JSON).")
        try:
            if "users" in st.secrets:
                secrets_val = st.secrets["users"]
                try:
                    _normalize_users(secrets_val)
                    diagnostics.append("st.secrets['users'] presente.")
                except Exception:
                    diagnostics.append("st.secrets['users'] presente pero inválido.")
        except Exception:
            pass

        st.error(
            "No hay usuarios configurados. Configura `CLV_USERS_JSON` o `CLV_USERS_B64`, "
            "`st.secrets['users']` o un archivo `users.json`."
        )
        if diagnostics:
            for d in diagnostics:
                st.warning(d)
        st.markdown("Ejemplo de `users.json`:")
        st.code('{\n  \"users\": {\n    \"admin\": {\"salt\": \"SALT\", \"hash\": \"SHA256\"}\n  }\n}', language="json")
        st.stop()

    if st.session_state.get("auth_user"):
        st.sidebar.success(f"Sesión: {st.session_state['auth_user']}")
        if st.sidebar.button("Cerrar sesión"):
            st.session_state["auth_user"] = None
            safe_rerun()
        return

    with st.sidebar.form("login_form"):
        st.markdown("### Acceso")
        username = st.text_input("Usuario")
        password = st.text_input("Clave", type="password")
        submitted = st.form_submit_button("Entrar")
        if submitted:
            if verify_user(username, password, users):
                st.session_state["auth_user"] = username
                safe_rerun()
            else:
                st.error("Usuario o clave incorrectos.")
    st.stop()

# --- Funciones de Cálculo de Informes (a implementar) ---

def detect_dayfirst(strings: pd.Series) -> bool:
    """Detecta si un formato ambiguo dd/mm o mm/dd es probablemente dayfirst."""
    if strings.empty:
        return True
    s = strings.astype(str).str.strip()
    mask = s.str.match(r"^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}$")
    if not mask.any():
        return True
    parts = s[mask].str.replace("-", "/", regex=False).str.split("/")
    day = pd.to_numeric(parts.str[0], errors="coerce")
    month = pd.to_numeric(parts.str[1], errors="coerce")
    dayfirst_votes = ((day > 12) & (month <= 12)).sum()
    monthfirst_votes = ((month > 12) & (day <= 12)).sum()
    if dayfirst_votes > monthfirst_votes:
        return True
    if monthfirst_votes > dayfirst_votes:
        return False
    return True

def detect_excel_origin(numeric: pd.Series) -> str:
    """Detecta si un serial Excel parece 1900 o 1904 según rango plausible."""
    if numeric.empty:
        return "1900"
    today = datetime.now()
    max_allowed = today + timedelta(days=365 * DATE_MAX_FUTURE_YEARS)
    min_allowed = DATE_MIN_ALLOWED
    origin_1900 = pd.Timestamp("1899-12-30")
    origin_1904 = pd.Timestamp("1904-01-01")
    dates_1900 = origin_1900 + pd.to_timedelta(numeric, unit="D")
    dates_1904 = origin_1904 + pd.to_timedelta(numeric, unit="D")
    score_1900 = ((dates_1900 >= min_allowed) & (dates_1900 <= max_allowed)).sum()
    score_1904 = ((dates_1904 >= min_allowed) & (dates_1904 <= max_allowed)).sum()
    return "1904" if score_1904 > score_1900 else "1900"

def parse_date_series(series: pd.Series, mode: str = "auto", logger=None) -> pd.Series:
    """Parsea una serie de fechas manejando Excel, ISO y formatos locales."""
    if logger is None:
        logger = lambda *args, **kwargs: None
    if pd.api.types.is_datetime64_any_dtype(series):
        return series

    s = series.copy()
    str_s = s.astype(str).str.strip()
    iso_mask = str_s.str.match(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}$")
    yyyymmdd_mask = str_s.str.match(r"^\d{8}$")

    def parse_excel(numeric, origin):
        base = pd.Timestamp("1899-12-30") if origin == "1900" else pd.Timestamp("1904-01-01")
        return base + pd.to_timedelta(numeric, unit="D")

    if mode == "excel_1900":
        numeric = pd.to_numeric(s, errors="coerce")
        logger("Formato forzado: Excel serial (1900)")
        return parse_excel(numeric, "1900")
    if mode == "excel_1904":
        numeric = pd.to_numeric(s, errors="coerce")
        logger("Formato forzado: Excel serial (1904)")
        return parse_excel(numeric, "1904")
    if mode == "iso":
        logger("Formato forzado: ISO (YYYY-MM-DD)")
        return pd.to_datetime(str_s, errors="coerce", dayfirst=False)
    if mode == "dmy":
        logger("Formato forzado: DD/MM/YYYY")
        result = pd.Series(pd.NaT, index=s.index)
        if iso_mask.any():
            result.loc[iso_mask] = pd.to_datetime(str_s[iso_mask], errors="coerce", dayfirst=False)
        result.loc[~iso_mask] = pd.to_datetime(str_s[~iso_mask], errors="coerce", dayfirst=True)
        return result
    if mode == "mdy":
        logger("Formato forzado: MM/DD/YYYY")
        result = pd.Series(pd.NaT, index=s.index)
        if iso_mask.any():
            result.loc[iso_mask] = pd.to_datetime(str_s[iso_mask], errors="coerce", dayfirst=False)
        result.loc[~iso_mask] = pd.to_datetime(str_s[~iso_mask], errors="coerce", dayfirst=False)
        return result
    if mode == "yyyymmdd":
        logger("Formato forzado: YYYYMMDD")
        return pd.to_datetime(str_s, errors="coerce", format="%Y%m%d")

    # --- Modo auto ---
    if yyyymmdd_mask.mean() > 0.7:
        logger("Formato detectado: YYYYMMDD")
        return pd.to_datetime(str_s, errors="coerce", format="%Y%m%d")

    numeric = pd.to_numeric(s, errors="coerce")
    numeric_ratio = numeric.notna().mean()

    # Si parece numérico, decidir entre YYYYMMDD o Excel
    if numeric_ratio > 0.7:
        numeric_nonnull = numeric[numeric.notna()]
        yyyymmdd_ratio_num = ((numeric_nonnull % 1 == 0) & numeric_nonnull.between(19000101, 21001231)).mean()
        if yyyymmdd_ratio_num > 0.7:
            logger("Formato detectado: YYYYMMDD (numérico)")
            result = pd.Series(pd.NaT, index=s.index)
            result.loc[numeric_nonnull.index] = pd.to_datetime(
                numeric_nonnull.astype(int).astype(str), errors="coerce", format="%Y%m%d"
            ).values
            rest_idx = s.index[~numeric.notna()]
            if len(rest_idx):
                result.loc[rest_idx] = parse_date_series(s.loc[rest_idx], mode="auto", logger=logger).values
            return result

        origin = detect_excel_origin(numeric_nonnull)
        logger(f"Formato detectado: Excel serial ({origin})")
        result = parse_excel(numeric, origin)
        rest_idx = s.index[~numeric.notna()]
        if len(rest_idx):
            result = result.copy()
            result.loc[rest_idx] = parse_date_series(s.loc[rest_idx], mode="auto", logger=logger).values
        return result

    if iso_mask.mean() > 0.6:
        logger("Formato detectado: ISO (YYYY-MM-DD)")
        return pd.to_datetime(str_s, errors="coerce", dayfirst=False)

    dayfirst = detect_dayfirst(str_s)
    logger(f"Formato detectado: {'DD/MM/YYYY' if dayfirst else 'MM/DD/YYYY'} (auto)")
    result = pd.Series(pd.NaT, index=s.index)
    if iso_mask.any():
        result.loc[iso_mask] = pd.to_datetime(str_s[iso_mask], errors="coerce", dayfirst=False)
    result.loc[~iso_mask] = pd.to_datetime(str_s[~iso_mask], errors="coerce", dayfirst=dayfirst)
    return result

def detect_amount_divisor(series: pd.Series, mode: str = "auto", logger=None) -> int:
    """Detecta si Total_pagado está en euros o céntimos para aplicar divisor."""
    if logger is None:
        logger = lambda *args, **kwargs: None
    if mode == "eur":
        logger("Total_pagado forzado: euros (divisor 1).")
        return 1
    if mode == "cents":
        logger("Total_pagado forzado: céntimos (divisor 100).")
        return 100
    numeric = pd.to_numeric(series, errors="coerce")
    numeric_nonnull = numeric[numeric.notna()]
    if numeric_nonnull.empty:
        return 1
    frac_ratio = ((numeric_nonnull % 1).abs() > 1e-6).mean()
    median_val = numeric_nonnull.median()
    max_val = numeric_nonnull.max()
    if frac_ratio > 0.3:
        logger("Total_pagado detectado con decimales: asumiendo euros (divisor 1).")
        return 1
    if median_val >= 1000 or max_val >= 100000:
        logger("Total_pagado detectado como valores altos: asumiendo céntimos (divisor 100).")
        return 100
    logger("Total_pagado detectado como valores moderados: asumiendo euros (divisor 1).")
    return 1

def preprocess_data(df, logger=st.write, date_mode="auto", amount_mode="auto"):
    """Limpia y preprocesa el DataFrame."""
    logger("Iniciando preprocesamiento de datos...")
    
    # Validar columnas
    required_columns = ['cod_cliente', 'Total_pagado']
    if not all(col in df.columns for col in required_columns):
        missing = [col for col in required_columns if col not in df.columns]
        logger(f"Faltan columnas requeridas: {', '.join(missing)}")
        return None
    if 'fecha' not in df.columns and 'fecha_hora' not in df.columns:
        logger("Falta columna de fecha: se requiere 'fecha' o 'fecha_hora'.")
        return None

    # Conversión de fechas
    logger("Convirtiendo fechas...")
    fecha_dt = None
    if 'fecha' in df.columns:
        fecha_dt = parse_date_series(df['fecha'], mode=date_mode, logger=logger)
    if fecha_dt is None:
        fecha_dt = pd.Series(pd.NaT, index=df.index)
    if 'fecha_hora' in df.columns:
        fecha_hora_dt = parse_date_series(df['fecha_hora'], mode="auto", logger=logger)
        missing_before = fecha_dt.isna().sum()
        fecha_dt = fecha_dt.fillna(fecha_hora_dt)
        filled = missing_before - fecha_dt.isna().sum()
        if filled > 0:
            logger(f"Se completaron {filled:,} fechas usando 'fecha_hora'.")
    df['fecha_dt'] = fecha_dt

    # Eliminar fechas inválidas o fuera de rango razonable
    before = len(df)
    df = df.dropna(subset=['fecha_dt'])
    min_allowed = DATE_MIN_ALLOWED
    max_allowed = datetime.now() + timedelta(days=365 * DATE_MAX_FUTURE_YEARS)
    out_of_range = (~df['fecha_dt'].between(min_allowed, max_allowed)).sum()
    if out_of_range > 0:
        logger(f"Se eliminaron {out_of_range:,} filas con fechas fuera de rango ({min_allowed.date()} a {max_allowed.date()}).")
    df = df[df['fecha_dt'].between(min_allowed, max_allowed)]
    logger(f"Filas después de limpiar fechas: {len(df):,} (antes: {before:,})")
    df = df.sort_values('fecha_dt')

    # Conversión de total pagado
    total_numeric = pd.to_numeric(df['Total_pagado'], errors='coerce')
    divisor = detect_amount_divisor(df['Total_pagado'], mode=amount_mode, logger=logger)
    df['Total_pagado_eur'] = total_numeric / divisor
    df = df.dropna(subset=['Total_pagado_eur'])
    logger(f"Filas después de limpiar importes: {len(df):,}")

    # Identificar primera compra
    logger("Identificando cohortes...")
    primera_compra = df.groupby('cod_cliente')['fecha_dt'].min().reset_index()
    primera_compra.columns = ['cod_cliente', 'primera_compra_dt']
    df = df.merge(primera_compra, on='cod_cliente')

    # Cohortes trimestrales y anuales
    df['cohorte_trimestral'] = df['primera_compra_dt'].dt.to_period('Q')
    df['periodo_trimestral'] = df['fecha_dt'].dt.to_period('Q')
    
    df['cohorte_anual'] = df['primera_compra_dt'].dt.to_period('Y')
    df['periodo_anual'] = df['fecha_dt'].dt.to_period('Y')

    logger("Preprocesamiento completado.")
    return df

def generate_retention_report(df):
    """Genera el informe de retención por trimestres."""
    retention_matrix = df.groupby(['cohorte_trimestral', 'periodo_trimestral'])['cod_cliente'].nunique().unstack(0)
    cohort_sizes = df.groupby('cohorte_trimestral')['cod_cliente'].nunique()
    
    retention_pct = retention_matrix.div(cohort_sizes, axis=1).T * 100
    retention_pct = retention_pct.fillna(0)  # mostrar churn explícito en vez de NaN
    
    # Formateo
    retention_pct.index = retention_pct.index.strftime('Y%Y-Q%q')
    retention_pct.columns = retention_pct.columns.strftime('Y%Y-Q%q')
    
    # Añadir total de clientes alineando índices para evitar errores de concatenación
    cohort_sizes.index = cohort_sizes.index.strftime('Y%Y-Q%q')
    cohort_sizes = cohort_sizes.reindex(retention_pct.index)  # asegurar alineación
    cohort_sizes = cohort_sizes.fillna(0)
    report_df = retention_pct.copy()
    report_df.insert(0, 'Total Clientes', cohort_sizes)

    return report_df

def generate_annual_retention_report(df):
    """Genera el informe de retención anual."""
    retention_matrix = df.groupby(['cohorte_anual', 'periodo_anual'])['cod_cliente'].nunique().unstack(0)
    cohort_sizes = df.groupby('cohorte_anual')['cod_cliente'].nunique()
    
    retention_pct = retention_matrix.div(cohort_sizes, axis=1).T * 100
    retention_pct = retention_pct.fillna(0)

    # Formateo
    retention_pct.index = retention_pct.index.strftime('%Y')
    retention_pct.columns = retention_pct.columns.strftime('%Y')

    cohort_sizes.index = cohort_sizes.index.strftime('%Y')
    cohort_sizes = cohort_sizes.reindex(retention_pct.index)  # asegurar alineación
    cohort_sizes = cohort_sizes.fillna(0)
    report_df = retention_pct.copy()
    report_df.insert(0, 'Total Clientes', cohort_sizes)

    return report_df

def generate_survival_analysis(df):
    """Genera el informe de análisis de supervivencia."""
    # Usar la fecha máxima del dataset como referencia temporal (evita depender del reloj actual)
    reference_date = df['fecha_dt'].max()
    if pd.isna(reference_date):
        reference_date = datetime.now()

    # Calcular meses desde la primera compra para cada pedido
    df['meses_desde_primera_compra'] = ((df['fecha_dt'].dt.year - df['primera_compra_dt'].dt.year) * 12 +
                                       (df['fecha_dt'].dt.month - df['primera_compra_dt'].dt.month))

    milestones = [0, 1, 3, 6, 9, 12, 18, 24, 36, 48, 60]
    
    # Pre-calcular el número de clientes únicos por cohorte
    cohort_sizes = df.groupby('cohorte_trimestral')['cod_cliente'].nunique()

    # Función para calcular supervivientes por cohorte
    def get_survivors_for_cohort(cohort_df):
        survivors = {}
        total_clients = cohort_df['cod_cliente'].nunique()
        for m in milestones:
            # Clientes que hicieron una compra EN O DESPUÉS del mes 'm'
            surviving_clients = cohort_df[cohort_df['meses_desde_primera_compra'] >= m]['cod_cliente'].nunique()
            survivors[f'Mes {m}'] = (surviving_clients / total_clients) * 100 if total_clients > 0 else 0
        return pd.Series(survivors)

    # Agrupar por cohorte y aplicar la función
    survival_table = df.groupby('cohorte_trimestral').apply(get_survivors_for_cohort)

    # Calcular métricas adicionales
    agg_stats = df.groupby('cod_cliente').agg(
        cohorte_trimestral=('cohorte_trimestral', 'first'),
        primera_compra=('primera_compra_dt', 'first'),
        ultima_compra=('fecha_dt', 'max'),
        total_pedidos=('cod_cliente', 'size'),
        total_revenue=('Total_pagado_eur', 'sum')
    ).reset_index()

    agg_stats['lifetime_dias'] = (agg_stats['ultima_compra'] - agg_stats['primera_compra']).dt.days
    
    cohort_stats = agg_stats.groupby('cohorte_trimestral').agg(
        Lifetime_Prom=('lifetime_dias', 'mean'),
        Pedidos_Prom=('total_pedidos', 'mean'),
        Revenue_Prom=('total_revenue', 'mean')
    ).reset_index()

    # Unir todo
    # Evitar duplicar columna 'cohorte_trimestral' si ya existe
    if 'cohorte_trimestral' in survival_table.columns:
        survival_table = survival_table.drop(columns=['cohorte_trimestral'])
    survival_table = survival_table.reset_index().rename(columns={'index': 'cohorte_trimestral'})
    report_df = pd.merge(survival_table, cohort_stats, on='cohorte_trimestral')
    
    # Añadir total de clientes y formatear
    report_df = pd.merge(report_df, cohort_sizes.rename('Total Clientes').reset_index(), on='cohorte_trimestral')
    report_df['cohorte_trimestral'] = report_df['cohorte_trimestral'].astype(str).str.replace('Q', '-Q')
    report_df = report_df.set_index('cohorte_trimestral')
    
    # Reordenar columnas
    cols = ['Total Clientes'] + [f'Mes {m}' for m in milestones] + ['Lifetime_Prom', 'Pedidos_Prom', 'Revenue_Prom']
    # Asegurar que todas las columnas existan; si falta alguna, crearla con NaN para evitar KeyError
    for col in cols:
        if col not in report_df.columns:
            report_df[col] = np.nan
    report_df = report_df[cols]
    
    # Calcular "Activos Hoy" (compras en los últimos 90 días)
    active_window_start = reference_date - timedelta(days=90)
    activos_hoy = df[df['fecha_dt'] >= active_window_start]['cod_cliente'].nunique()
    total_clientes = df['cod_cliente'].nunique()
    one_time_buyers = agg_stats[agg_stats['total_pedidos'] == 1]['cod_cliente'].nunique()

    summary = {
        "Total clientes analizados": total_clientes,
        "Clientes activos (últimos 90 días)": activos_hoy,
        "Tiempo de vida promedio (días)": agg_stats['lifetime_dias'].mean(),
        "Promedio de pedidos por cliente": agg_stats['total_pedidos'].mean(),
        "% de clientes con 1 sola compra": (one_time_buyers / total_clientes) * 100 if total_clientes > 0 else 0
    }

    active_window = {
        "start": active_window_start.date(),
        "end": reference_date.date()
    }
    
    return report_df, summary, active_window


def generate_frequency_report(df):
    """Genera el informe de frecuencia de compra con sus 4 secciones."""
    # Base para el análisis: clientes con 2+ compras
    df_sorted = df.sort_values(['cod_cliente', 'fecha_dt'])
    df_sorted['dias_desde_anterior'] = df_sorted.groupby('cod_cliente')['fecha_dt'].diff().dt.days
    clientes_multi_compra = df_sorted.dropna(subset=['dias_desde_anterior'])
    
    # --- 1. Distribución por Frecuencia de Compra ---
    bins = [0, 30, 60, 90, 180, 365, np.inf]
    labels = ['≤30 días', '31-60 días', '61-90 días', '91-180 días', '181-365 días', '>365 días']
    clientes_multi_compra['segmento_frecuencia'] = pd.cut(clientes_multi_compra['dias_desde_anterior'], bins=bins, labels=labels, right=True)
    
    distribucion_frecuencia = clientes_multi_compra.groupby('segmento_frecuencia').agg(
        Total_Intervalos=('dias_desde_anterior', 'count'),
        Dias_Promedio=('dias_desde_anterior', 'mean'),
        Dias_Mediana=('dias_desde_anterior', 'median')
    )
    total_intervalos = distribucion_frecuencia['Total_Intervalos'].sum()
    distribucion_frecuencia['% del Total'] = (distribucion_frecuencia['Total_Intervalos'] / total_intervalos) * 100
    distribucion_frecuencia = distribucion_frecuencia[['Total_Intervalos', '% del Total', 'Dias_Promedio', 'Dias_Mediana']]

    # --- 2. Tiempo hasta Segunda Compra ---
    df_sorted['num_compra'] = df_sorted.groupby('cod_cliente').cumcount() + 1
    compras_1_y_2 = df_sorted[df_sorted['num_compra'].isin([1, 2])]
    
    tiempo_segunda_compra_raw = compras_1_y_2.groupby('cod_cliente').agg(
        num_compras=('num_compra', 'max'),
        dias_hasta_segunda=('dias_desde_anterior', 'last')
    )
    tiempo_segunda_compra_raw = tiempo_segunda_compra_raw[tiempo_segunda_compra_raw['num_compras'] > 1]
    
    bins_2da = [-1, 30, 60, 90, 180, np.inf]
    labels_2da = ['Dentro de 30 días', '31-60 días', '61-90 días', '91-180 días', 'Más de 180 días']
    tiempo_segunda_compra_raw['periodo'] = pd.cut(tiempo_segunda_compra_raw['dias_hasta_segunda'], bins=bins_2da, labels=labels_2da, right=True)
    
    tiempo_segunda_compra = tiempo_segunda_compra_raw.groupby('periodo').agg(
        Clientes=('dias_hasta_segunda', 'count')
    )
    total_clientes_2da = tiempo_segunda_compra['Clientes'].sum()
    tiempo_segunda_compra['% del Total'] = (tiempo_segunda_compra['Clientes'] / total_clientes_2da) * 100

    # --- 3. Evolución de Frecuencia por Número de Compra ---
    max_compras = 10 # Limitar para legibilidad
    evolucion = df_sorted[df_sorted['num_compra'] <= max_compras]
    
    evolucion_frecuencia = evolucion.groupby('num_compra').agg(
        Numero_Clientes=('cod_cliente', 'nunique'),
        Dias_Promedio_Intervalo=('dias_desde_anterior', 'mean'),
        Dias_Mediana_Intervalo=('dias_desde_anterior', 'median')
    ).reset_index()
    evolucion_frecuencia = evolucion_frecuencia[evolucion_frecuencia['num_compra'] > 1] # El intervalo es para la compra N
    evolucion_frecuencia['Tendencia'] = evolucion_frecuencia['Dias_Promedio_Intervalo'].diff().apply(
        lambda x: '↓ Mejora' if x < 0 else ('↑ Empeora' if x > 0 else '→ Estable')
    )

    # --- 4. Velocidad de Compra (Compras por Mes) ---
    agg_stats = df.groupby('cod_cliente').agg(
        total_pedidos=('cod_cliente', 'size'),
        primera_compra=('primera_compra_dt', 'first'),
        ultima_compra=('fecha_dt', 'max'),
        total_revenue=('Total_pagado_eur', 'sum')
    ).reset_index()
    
    agg_stats = agg_stats[agg_stats['total_pedidos'] > 1]
    agg_stats['dias_actividad'] = (agg_stats['ultima_compra'] - agg_stats['primera_compra']).dt.days
    # Si primera y última compra son el mismo día, pero hay >1 pedido, considerar 1 día de actividad.
    agg_stats.loc[agg_stats['dias_actividad'] == 0, 'dias_actividad'] = 1
    
    agg_stats['compras_por_mes'] = agg_stats['total_pedidos'] / (agg_stats['dias_actividad'] / 30)

    bins_vel = [-np.inf, 0.1, 0.25, 0.5, 1, np.inf]
    labels_vel = ['Muy Baja (<0.1)', 'Baja (0.1-0.24)', 'Media (0.25-0.49)', 'Media-Alta (0.5-0.99)', 'Alta (≥1)']
    agg_stats['segmento_velocidad'] = pd.cut(agg_stats['compras_por_mes'], bins=bins_vel, labels=labels_vel, right=False)

    velocidad_compra = agg_stats.groupby('segmento_velocidad').agg(
        Numero_Clientes=('cod_cliente', 'count'),
        Compras_por_Mes_Promedio=('compras_por_mes', 'mean'),
        Pedidos_Promedio=('total_pedidos', 'mean'),
        Revenue_Promedio=('total_revenue', 'mean')
    )
    total_clientes_vel = velocidad_compra['Numero_Clientes'].sum()
    velocidad_compra['% del Total'] = (velocidad_compra['Numero_Clientes'] / total_clientes_vel) * 100

    return {
        "distribucion": distribucion_frecuencia,
        "segunda_compra": tiempo_segunda_compra,
        "evolucion": evolucion_frecuencia.set_index('num_compra'),
        "velocidad": velocidad_compra
    }

def generate_loyalty_report(df):
    """Genera métricas de activación, repetición e ingresos recurrentes por año."""
    if df is None or df.empty or 'fecha_dt' not in df.columns:
        return None
    data = df.copy()
    data['year'] = data['fecha_dt'].dt.year
    if 'Num_Pedidos_Cliente' in data.columns:
        data['Num_Pedidos_Cliente'] = pd.to_numeric(data['Num_Pedidos_Cliente'], errors='coerce')
    if 'cliente_nuevo' in data.columns:
        data['cliente_nuevo'] = data['cliente_nuevo'].astype(str).str.lower()

    grp = data.groupby('year')

    # Tasa de activación
    if 'Num_Pedidos_Cliente' in data.columns:
        act_num = grp.apply(lambda g: g.loc[g['Num_Pedidos_Cliente'] == 2, 'cod_cliente'].nunique())
        act_den = grp.apply(lambda g: g.loc[g['Num_Pedidos_Cliente'] == 1, 'cod_cliente'].nunique())
        tasa_activacion = (act_num / act_den * 100).replace([np.inf, -np.inf], np.nan)
    else:
        tasa_activacion = pd.Series(index=grp.size().index, dtype=float)

    # Tasa de repetición
    email_col = 'email' if 'email' in data.columns else 'cod_cliente'
    if 'cliente_nuevo' in data.columns:
        rep_num = grp.apply(lambda g: g.loc[g['cliente_nuevo'] == 'n', email_col].nunique())
    else:
        rep_num = pd.Series(index=grp.size().index, dtype=float)
    rep_den = grp.apply(lambda g: g[email_col].nunique())
    tasa_repeticion = (rep_num / rep_den * 100).replace([np.inf, -np.inf], np.nan)

    # Tasa de ingresos de clientes recurrentes
    if 'cliente_nuevo' in data.columns and 'Total_pagado_eur' in data.columns:
        rev_num = grp.apply(lambda g: g.loc[g['cliente_nuevo'] == 'n', 'Total_pagado_eur'].sum())
        rev_den = grp.apply(lambda g: g['Total_pagado_eur'].sum())
        tasa_ingresos = (rev_num / rev_den * 100).replace([np.inf, -np.inf], np.nan)
    else:
        tasa_ingresos = pd.Series(index=grp.size().index, dtype=float)

    # Total pedidos
    if 'codigo' in data.columns:
        total_pedidos = grp['codigo'].nunique()
    else:
        total_pedidos = grp.size()

    df_activacion = pd.DataFrame({'Tasa_activacion': tasa_activacion}).sort_index()
    df_repeticion = pd.DataFrame({'Tasa_repeticion': tasa_repeticion}).sort_index()
    df_ingresos = pd.DataFrame({'Tasa_ingresos_recurrentes': tasa_ingresos}).sort_index()
    df_rep_ped = pd.DataFrame({'Tasa_repeticion': tasa_repeticion, 'Total_pedidos': total_pedidos}).sort_index()

    return {
        "activacion": df_activacion,
        "repeticion": df_repeticion,
        "ingresos_recurrentes": df_ingresos,
        "repeticion_pedidos": df_rep_ped
    }


# Las funciones para los informes 3 y 4 son más complejas y se añadirán progresivamente.

def style_retention_table(df):
    """Aplica estilos de color a la tabla de retención."""
    if df is None or df.empty:
        return df

    def build_styles(data):
        styles = pd.DataFrame("", index=data.index, columns=data.columns)
        if len(data.columns) <= 1:
            return styles
        for i, row_label in enumerate(data.index):
            for j, col_label in enumerate(data.columns[1:], start=1):
                val = data.iloc[i, j]
                # Convertir 'Y2024-Q1' a un objeto Period
                try:
                    cohorte_period = pd.Period(str(row_label).replace('Y', ''), freq='Q')
                    current_period = pd.Period(str(col_label).replace('Y', ''), freq='Q')
                except Exception:
                    continue

                if pd.isna(val) or val == 0 or current_period < cohorte_period:
                    styles.iloc[i, j] = 'background-color: #C0C0C0' # Gris
                elif current_period == cohorte_period:
                    styles.iloc[i, j] = 'background-color: #1E6B1E; color: white' # Verde oscuro
                elif val >= 8:
                    styles.iloc[i, j] = 'background-color: #7CCD7C' # Verde claro
                elif 3 <= val < 8:
                    styles.iloc[i, j] = 'background-color: #FFD700' # Amarillo
                else:
                    styles.iloc[i, j] = 'background-color: #FF6B6B' # Rojo
        return styles

    first_col = df.columns[0] if len(df.columns) else None
    fmt = {}
    if first_col:
        fmt[first_col] = "{:,.0f}"
    for col in df.columns[1:]:
        fmt[col] = "{:.2f}%"

    styled = df.style.format(fmt, na_rep="").apply(
        build_styles,
        axis=None
    )
    return styled

def style_heatmap(df, cmap="Greens"):
    """Aplica un gradiente de color a valores numéricos, mantiene la primera col sin formato."""
    if df is None or df.empty:
        return df
    first_col = df.columns[0]
    fmt = {}
    if first_col in df.columns:
        fmt[first_col] = "{:,.0f}"
    for col in df.columns[1:]:
        fmt[col] = "{:.2f}"
    styled = df.style.format(fmt, na_rep="-")
    
    # Filter to apply gradient only on numeric columns from the second column onwards
    numeric_cols = [col for col in df.columns[1:] if pd.api.types.is_numeric_dtype(df[col])]
    
    if numeric_cols:
        styled = styled.background_gradient(axis=None, cmap=cmap, subset=numeric_cols)

    return styled

def style_percent_heatmap(df, cmap="Blues"):
    """Gradiente para porcentajes: rojo solo en 0; resto verde claro→oscuro."""
    if df is None or df.empty:
        return df
    styled = df.style
    fmt = {}
    if df.columns[0].lower().startswith("total"):
        fmt[df.columns[0]] = "{:,.0f}"
        percent_cols = df.columns[1:]
    else:
        percent_cols = df.columns
    for col in percent_cols:
        fmt[col] = "{:.2f}%"
    styled = styled.format(fmt, na_rep="-")

    # Color mapping: 0 -> rojo, >0 -> gradiente verde
    def color_map(val):
        if pd.isna(val):
            return ""
        if val == 0:
            return "background-color: #ff6b6b; color: white"
        # Interpolar verde claro (#e8f5e9) a verde oscuro (#1b5e20)
        v = max(0.0, min(100.0, float(val)))
        t = v / 100.0
        start = np.array([0xE8, 0xF5, 0xE9])
        end = np.array([0x1B, 0x5E, 0x20])
        rgb = (start + (end - start) * t).astype(int)
        return f"background-color: rgb({rgb[0]},{rgb[1]},{rgb[2]}); color: {'white' if t>0.55 else 'black'}"

    styled = styled.applymap(color_map, subset=percent_cols)
    return styled

def style_currency(df, cols):
    if df is None or df.empty:
        return df
    fmt = {col: "€{:.2f}" for col in cols if col in df.columns}
    return df.style.format(fmt)

def style_frequency_distribucion(df):
    if df is None or df.empty:
        return df
    styled = df.style.format({
        'Total_Intervalos': "{:,.0f}",
        '% del Total': "{:.2f}%",
        'Dias_Promedio': "{:.1f}",
        'Dias_Mediana': "{:.1f}",
    })
    if '% del Total' in df.columns:
        styled = styled.background_gradient(cmap=RETENTION_CMAP, subset=['% del Total'])
    day_cols = [c for c in ['Dias_Promedio', 'Dias_Mediana'] if c in df.columns]
    if day_cols:
        styled = styled.background_gradient(cmap=FREQ_DISTRIB_CMAP, subset=day_cols)
    return styled

def style_frequency_second_purchase(df):
    if df is None or df.empty:
        return df
    styled = df.style.format({
        'Clientes': "{:,.0f}",
        '% del Total': "{:.2f}%"
    })
    if '% del Total' in df.columns:
        styled = styled.background_gradient(cmap=RETENTION_CMAP, subset=['% del Total'])
    return styled

def style_frequency_evolucion(df):
    if df is None or df.empty:
        return df
    styled = df.style.format({
        'Numero_Clientes': "{:,.0f}",
        'Dias_Promedio_Intervalo': "{:.1f}",
        'Dias_Mediana_Intervalo': "{:.1f}",
    })
    day_cols = [c for c in ['Dias_Promedio_Intervalo', 'Dias_Mediana_Intervalo'] if c in df.columns]
    if day_cols:
        styled = styled.background_gradient(cmap=FREQ_EVOL_CMAP, subset=day_cols)
    return styled

def style_frequency_velocidad(df):
    if df is None or df.empty:
        return df
    styled = df.style.format({
        'Numero_Clientes': "{:,.0f}",
        'Compras_por_Mes_Promedio': "{:.2f}",
        'Pedidos_Promedio': "{:.2f}",
        'Revenue_Promedio': "€{:.2f}",
        '% del Total': "{:.2f}%"
    })
    grad_cols = [c for c in ['Compras_por_Mes_Promedio', 'Pedidos_Promedio', 'Revenue_Promedio'] if c in df.columns]
    if grad_cols:
        styled = styled.background_gradient(cmap=FREQ_VEL_CMAP, subset=grad_cols)
    if '% del Total' in df.columns:
        styled = styled.background_gradient(cmap=RETENTION_CMAP, subset=['% del Total'])
    return styled

def style_repetition_orders(df):
    if df is None or df.empty:
        return df
    fmt = {}
    if 'Tasa_repeticion' in df.columns:
        fmt['Tasa_repeticion'] = "{:.2f}%"
    if 'Total_pedidos' in df.columns:
        fmt['Total_pedidos'] = "{:,.0f}"
    styled = df.style.format(fmt, na_rep="-")
    if 'Tasa_repeticion' in df.columns:
        styled = styled.background_gradient(cmap=RETENTION_CMAP, subset=['Tasa_repeticion'])
    return styled

def plot_repetition_vs_orders(df):
    """Grafico de doble eje: tasa de repetición (%) y total pedidos."""
    if df is None or df.empty:
        return None
    if 'Tasa_repeticion' not in df.columns or 'Total_pedidos' not in df.columns:
        return None
    try:
        import matplotlib.pyplot as plt
    except Exception:
        return None

    years = df.index.astype(str).tolist()
    x = np.arange(len(years))
    width = 0.38
    tasa = df['Tasa_repeticion'].astype(float).values
    pedidos = df['Total_pedidos'].astype(float).values

    fig, ax1 = plt.subplots(figsize=(8, 3.5))
    ax2 = ax1.twinx()

    bars1 = ax1.bar(x - width/2, tasa, width, color="#4C8BF5", label="Tasa de repetición (%)")
    bars2 = ax2.bar(x + width/2, pedidos, width, color="#F4A261", label="Total pedidos")

    ax1.set_ylabel("% repetición")
    ax2.set_ylabel("Total pedidos")
    ax1.set_xticks(x)
    ax1.set_xticklabels(years)
    ax1.set_ylim(0, max(100, np.nanmax(tasa) * 1.2 if len(tasa) else 100))
    ax1.grid(axis="y", linestyle="--", alpha=0.3)

    # Etiquetas de valores
    for b in bars1:
        val = b.get_height()
        ax1.annotate(
            f"{val:.2f}%",
            xy=(b.get_x() + b.get_width()/2, val),
            xytext=(0, 3),
            textcoords="offset points",
            ha="center",
            va="bottom",
            fontsize=8,
            color="#1F3B75",
        )
    for b in bars2:
        val = b.get_height()
        ax2.annotate(
            f"{val:,.0f}",
            xy=(b.get_x() + b.get_width()/2, val),
            xytext=(0, 3),
            textcoords="offset points",
            ha="center",
            va="bottom",
            fontsize=8,
            color="#7A3E00",
        )

    # Leyenda combinada
    h1, l1 = ax1.get_legend_handles_labels()
    h2, l2 = ax2.get_legend_handles_labels()
    ax1.legend(h1 + h2, l1 + l2, loc="upper left", frameon=False)

    fig.tight_layout()
    return fig

def plot_bar_with_labels(df, y_col, color="#4C8BF5", is_percent=True):
    """Gráfico de barras con etiquetas de valor encima."""
    if df is None or df.empty or y_col not in df.columns:
        return None
    try:
        import matplotlib.pyplot as plt
    except Exception:
        return None

    years = df.index.astype(str).tolist()
    x = np.arange(len(years))
    values = pd.to_numeric(df[y_col], errors="coerce").values

    fig, ax = plt.subplots(figsize=(8, 3.5))
    bars = ax.bar(x, values, color=color)

    ax.set_xticks(x)
    ax.set_xticklabels(years)
    ax.grid(axis="y", linestyle="--", alpha=0.3)
    if is_percent:
        ax.set_ylim(0, max(100, np.nanmax(values) * 1.2 if len(values) else 100))
        ax.set_ylabel("%")
    else:
        ax.set_ylabel("Valor")

    for b in bars:
        val = b.get_height()
        if pd.isna(val):
            continue
        label = f"{val:.2f}%" if is_percent else f"{val:,.0f}"
        ax.annotate(
            label,
            xy=(b.get_x() + b.get_width()/2, val),
            xytext=(0, 3),
            textcoords="offset points",
            ha="center",
            va="bottom",
            fontsize=8,
            color="#1F3B75",
        )

    fig.tight_layout()
    return fig

def style_survival_table(df):
    """Estilos para tabla de supervivencia: % en Meses, días/pedidos/revenue en columnas finales."""
    if df is None or df.empty:
        return df
    data = df.copy()
    # Limpiar posibles strings con % o € en columnas finales
    for col in ["Lifetime_Prom", "Pedidos_Prom", "Revenue_Prom"]:
        if col in data.columns and data[col].dtype == object:
            data[col] = (
                data[col]
                .astype(str)
                .str.replace("%", "", regex=False)
                .str.replace("€", "", regex=False)
            )
            data[col] = pd.to_numeric(data[col], errors="coerce")

    cols = data.columns
    fmt = {}
    if len(cols):
        if cols[0].lower().startswith("total"):
            fmt[cols[0]] = "{:,.0f}"
    mes_cols = [c for c in cols if str(c).startswith("Mes ")]
    for c in mes_cols:
        fmt[c] = "{:.2f}%"
    if "Lifetime_Prom" in cols:
        fmt["Lifetime_Prom"] = "{:.1f}"
    if "Pedidos_Prom" in cols:
        fmt["Pedidos_Prom"] = "{:.2f}"
    if "Revenue_Prom" in cols:
        fmt["Revenue_Prom"] = "€{:.2f}"

    styled = data.style.format(fmt, na_rep="-")
    if mes_cols:
        scale_cols = [c for c in mes_cols if str(c) != "Mes 0"]
        if not scale_cols:
            scale_cols = mes_cols
        max_val = pd.to_numeric(data[scale_cols].stack(), errors="coerce").max()
        if pd.isna(max_val) or max_val <= 0:
            max_val = 1.0

        def color_map(val):
            if pd.isna(val):
                return ""
            if val == 0:
                return "background-color: #ff6b6b; color: white"
            # Interpolar verde claro (#e8f5e9) a verde oscuro (#1b5e20)
            t = max(0.0, min(1.0, float(val) / max_val))
            start = np.array([0xE8, 0xF5, 0xE9])
            end = np.array([0x1B, 0x5E, 0x20])
            rgb = (start + (end - start) * t).astype(int)
            return f"background-color: rgb({rgb[0]},{rgb[1]},{rgb[2]}); color: {'white' if t>0.55 else 'black'}"

        styled = styled.applymap(color_map, subset=mes_cols)
    if "Lifetime_Prom" in cols:
        styled = styled.background_gradient(cmap="Greens", subset=["Lifetime_Prom"])
    return styled

def styler_supported():
    """Devuelve True si la versión de Streamlit permite renderizar pandas.Styler en st.dataframe."""
    try:
        major, minor, *_ = map(int, st.__version__.split('.')[:2])
        return (major > 1) or (major == 1 and minor >= 31)
    except Exception:
        return False

def show_table(df, styler_fn=None, info_msg=None):
    """Renderiza df con estilos si la versión de Streamlit lo soporta; si no, muestra fallback plano."""
    if styler_supported() and styler_fn is not None:
        st.dataframe(styler_fn(df))
    else:
        if info_msg:
            st.info(info_msg)
        st.dataframe(df)

def style_delta_retention(df):
    """Estilos para diferencias (B - A) en retención: verde positivo, rojo negativo."""
    if df is None or df.empty:
        return df
    first_col = df.columns[0] if len(df.columns) else None
    styled = df.style
    fmt = {}
    if first_col:
        fmt[first_col] = "{:+,.0f}"
        percent_cols = df.columns[1:]
    else:
        percent_cols = df.columns
    for col in percent_cols:
        fmt[col] = "{:+.2f}%"
    styled = styled.format(fmt, na_rep="-")

    def color_sign(val):
        if pd.isna(val):
            return ""
        if val > 0:
            return "background-color: #c8e6c9; color: #1b5e20"
        if val < 0:
            return "background-color: #ffcdd2; color: #b71c1c"
        return "background-color: #eeeeee; color: #424242"

    if first_col:
        styled = styled.applymap(color_sign, subset=[first_col])
    if len(percent_cols):
        styled = styled.applymap(color_sign, subset=percent_cols)
    return styled

def style_delta_counts(df):
    """Estilos para diferencias en conteos: verde positivo, rojo negativo."""
    if df is None or df.empty:
        return df
    styled = df.style.format("{:+,.0f}", na_rep="-")

    def color_sign(val):
        if pd.isna(val):
            return ""
        if val > 0:
            return "background-color: #c8e6c9; color: #1b5e20"
        if val < 0:
            return "background-color: #ffcdd2; color: #b71c1c"
        return "background-color: #eeeeee; color: #424242"

    styled = styled.applymap(color_sign)
    return styled

def normalize_retention_df(df):
    """Normaliza índices/columnas y fuerza columnas numéricas en reportes de retención."""
    if df is None or df.empty:
        return df
    df_norm = df.copy()
    df_norm.index = df_norm.index.map(str)
    df_norm.columns = df_norm.columns.map(str)
    if len(df_norm.columns):
        first_col = df_norm.columns[0]
        df_norm[first_col] = pd.to_numeric(df_norm[first_col], errors="coerce")
        for col in df_norm.columns[1:]:
            df_norm[col] = pd.to_numeric(df_norm[col], errors="coerce")
    return df_norm

def sort_period_labels(labels, freq):
    """Ordena etiquetas de periodos (Q/Y) de forma cronológica si es posible."""
    def _key(x):
        try:
            if freq == "Q":
                return pd.Period(str(x).replace("Y", ""), freq="Q").start_time
            if freq == "Y":
                return pd.Period(str(x).replace("Y", ""), freq="Y").start_time
        except Exception:
            return pd.Timestamp.min
        return pd.Timestamp.min
    return sorted(set(labels), key=_key)

def align_retention_tables(df_a, df_b, freq):
    """Alinea dos tablas de retención por índice y columnas."""
    df_a = normalize_retention_df(df_a)
    df_b = normalize_retention_df(df_b)
    if df_a is None or df_b is None:
        return df_a, df_b
    idx = sort_period_labels(list(df_a.index) + list(df_b.index), freq=freq)

    # Mantener la primera columna de totales y ordenar el resto por periodo
    first_col = df_a.columns[0] if len(df_a.columns) else (df_b.columns[0] if len(df_b.columns) else "Total Clientes")
    other_cols = [c for c in list(df_a.columns) + list(df_b.columns) if c != first_col]
    other_cols = sort_period_labels(other_cols, freq=freq)
    cols = [first_col] + [c for c in other_cols if c != first_col]

    df_a = df_a.reindex(index=idx, columns=cols)
    df_b = df_b.reindex(index=idx, columns=cols)
    return df_a, df_b

def retention_counts_from_pct(df):
    """Calcula conteos estimados por celda (pct * total clientes)."""
    if df is None or df.empty:
        return df
    df_counts = df.copy()
    total_col = df_counts.columns[0]
    for col in df_counts.columns[1:]:
        df_counts[col] = (df_counts[col] / 100.0 * df_counts[total_col]).round(0)
    return df_counts

def filter_reports_by_date(reports, start_date, end_date):
    """Filtra tablas de informes solo para visualización según rango de fechas sin recalcular CSV."""
    if not reports:
        return None
    start_ts = pd.Timestamp(start_date)
    end_ts = pd.Timestamp(end_date) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)

    def filter_period_table(df, freq):
        if df is None or df.empty:
            return df
        df_copy = df.copy()
        # Filtrar filas (cohortes)
        idx_period = df_copy.index.to_series().apply(lambda x: pd.Period(str(x).replace('Y',''), freq=freq))
        row_mask = (idx_period.dt.start_time >= start_ts) & (idx_period.dt.start_time <= end_ts)
        df_copy = df_copy.loc[row_mask]
        # Filtrar columnas (periodos calendario)
        # Inicializar col_keep con el mismo índice y longitud que df_copy.columns
        col_keep = pd.Series(False, index=df_copy.columns)
        # Siempre mantener la primera columna de totales (índice 0)
        if not df_copy.columns.empty:
            col_keep.iloc[0] = True

        # Crear col_period solo para las columnas relevantes (a partir de la segunda)
        if len(df_copy.columns) > 1:
            col_period_names = df_copy.columns[1:]
            col_period_series = pd.Series(col_period_names, index=col_period_names).apply(lambda x: pd.Period(str(x).replace('Y',''), freq=freq))
            period_mask = (col_period_series.dt.start_time >= start_ts) & (col_period_series.dt.start_time <= end_ts)
            col_keep[1:] = period_mask.values # Asignar los valores booleanos a las columnas correspondientes
        
        df_copy = df_copy.loc[:, col_keep]
        return df_copy

    def filter_year_table(df):
        if df is None or df.empty:
            return df
        df_copy = df.copy()
        years = pd.to_numeric(df_copy.index, errors="coerce")
        mask = (years >= start_ts.year) & (years <= end_ts.year)
        return df_copy.loc[mask]

    filtered = {}
    filtered['report1'] = filter_period_table(reports['report1'], 'Q') if 'report1' in reports else None
    filtered['report2'] = filter_period_table(reports['report2'], 'Y') if 'report2' in reports else None
    # Supervivencia: filtrar filas por cohorte
    if 'report3' in reports and reports['report3'] is not None:
        df3 = reports['report3'].copy()
        idx_period = df3.index.to_series().apply(lambda x: pd.Period(str(x).replace('Q','-Q') if 'Q' not in str(x) else str(x), freq='Q'))
        row_mask = (idx_period.dt.start_time >= start_ts) & (idx_period.dt.start_time <= end_ts)
        df3 = df3.loc[row_mask]
        filtered['report3'] = df3
    else:
        filtered['report3'] = None
    filtered['report4'] = reports.get('report4')
    if 'report5' in reports and reports['report5'] is not None:
        r5 = reports['report5']
        filtered['report5'] = {
            "activacion": filter_year_table(r5.get('activacion')),
            "repeticion": filter_year_table(r5.get('repeticion')),
            "ingresos_recurrentes": filter_year_table(r5.get('ingresos_recurrentes')),
            "repeticion_pedidos": filter_year_table(r5.get('repeticion_pedidos')),
        }
    else:
        filtered['report5'] = None
    return filtered

from openpyxl.styles import PatternFill, Font

def export_to_excel(reports):
    """Crea un archivo Excel en memoria con los informes y sus estilos."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Escribir cada informe en una hoja
        reports['report1'].to_excel(writer, sheet_name='Retención Trimestral')
        reports['report2'].to_excel(writer, sheet_name='Retención Anual')
        reports['report3'].to_excel(writer, sheet_name='Análisis de Supervivencia')
        
        # Informes de frecuencia
        reports['report4']['distribucion'].to_excel(writer, sheet_name='Frecuencia - Distribución')
        reports['report4']['segunda_compra'].to_excel(writer, sheet_name='Frecuencia - 2da Compra')
        reports['report4']['evolucion'].to_excel(writer, sheet_name='Frecuencia - Evolución')
        reports['report4']['velocidad'].to_excel(writer, sheet_name='Frecuencia - Velocidad')

        # Informes de lealtad (frecuencia y recurrencia)
        if reports.get('report5'):
            r5 = reports['report5']
            if r5.get('activacion') is not None:
                r5['activacion'].to_excel(writer, sheet_name='Lealtad - Activación')
            if r5.get('repeticion') is not None:
                r5['repeticion'].to_excel(writer, sheet_name='Lealtad - Repetición')
            if r5.get('ingresos_recurrentes') is not None:
                r5['ingresos_recurrentes'].to_excel(writer, sheet_name='Lealtad - Ingresos Recurrentes')
            if r5.get('repeticion_pedidos') is not None:
                r5['repeticion_pedidos'].to_excel(writer, sheet_name='Lealtad - Repetición vs Pedidos')

        # --- Aplicar Estilos (Ejemplo para Retención Trimestral) ---
        # Definir rellenos de color
        green_dark_fill = PatternFill(start_color="1E6B1E", end_color="1E6B1E", fill_type="solid")
        green_light_fill = PatternFill(start_color="7CCD7C", end_color="7CCD7C", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        gray_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        white_font = Font(color="FFFFFF")

        ws = writer.sheets['Retención Trimestral']
        df = reports['report1']

        # Iterar sobre las celdas de datos para aplicar el formato condicional
        for r_idx, row in enumerate(df.index, 2): # 2 porque Excel es 1-based y hay cabecera
            for c_idx, col_name in enumerate(df.columns, 2): # Empieza en la 3a col del excel
                cell = ws.cell(row=r_idx, column=c_idx)
                val = df.loc[row, col_name]

                if pd.isna(val):
                    continue

                try:
                    cohorte_period = pd.Period(row.replace('Y', ''), freq='Q')
                    current_period = pd.Period(col_name.replace('Y', ''), freq='Q')
                    
                    if current_period < cohorte_period:
                        cell.fill = gray_fill
                    elif current_period == cohorte_period:
                        cell.fill = green_dark_fill
                        cell.font = white_font
                    elif val >= 8:
                        cell.fill = green_light_fill
                    elif 3 <= val < 8:
                        cell.fill = yellow_fill
                    elif val > 0: # Solo colorear si hay retención
                        cell.fill = red_fill

                except Exception as e:
                    # Ignorar errores de parseo de fechas para columnas no periódicas
                    pass
    
    output.seek(0)
    return output


# --- UI de la Aplicación Streamlit ---

st.set_page_config(layout="wide", page_title="Análisis CLV Ecommerce")

# Paletas de color más contrastadas
RETENTION_CMAP = "RdYlGn"
SURVIVAL_CMAP = "YlGn"
FREQ_DISTRIB_CMAP = "BuGn"
FREQ_EVOL_CMAP = "PuBu"
FREQ_VEL_CMAP = "OrRd"

if "df_raw" not in st.session_state: # Inicializar df_raw también
    st.session_state["df_raw"] = None
if "selected_report" not in st.session_state: # Para almacenar el nombre del informe cargado/generado
    st.session_state["selected_report"] = None
if "last_generated_excel_bytes" not in st.session_state:
    st.session_state["last_generated_excel_bytes"] = None
if "last_generated_reports" not in st.session_state:
    st.session_state["last_generated_reports"] = None
if "last_generated_summary" not in st.session_state:
    st.session_state["last_generated_summary"] = None
if "last_generated_range" not in st.session_state:
    st.session_state["last_generated_range"] = None
if "last_generated_active_window" not in st.session_state:
    st.session_state["last_generated_active_window"] = None
if "base_active_window" not in st.session_state:
    st.session_state["base_active_window"] = None
if "delete_candidate" not in st.session_state:
    st.session_state["delete_candidate"] = None
if "nav_to" not in st.session_state:
    st.session_state["nav_to"] = None
if "auth_user" not in st.session_state:
    st.session_state["auth_user"] = None

preview_placeholder = st.empty()
status_placeholder = st.empty()

# Control de acceso (bloquea el resto de la app hasta autenticación)
require_auth()

# Navegación principal
st.sidebar.markdown("---")
st.sidebar.markdown("### Opciones Principales")
if st.session_state.get("nav_to"):
    st.session_state["app_mode"] = st.session_state["nav_to"]
    st.session_state["nav_to"] = None
mode = st.sidebar.radio(
    "¿Qué deseas hacer?",
    ("Generar informe", "Informes guardados", "Ver informe", "Comparar informes"),
    index=2 if st.session_state.get("base_reports") else 0,
    key="app_mode"
)

# Encabezado principal (compacto en modo ver informe)
if mode != "Ver informe":
    st.title("Análisis CLV Ecommerce")
    st.markdown("Sube un archivo CSV de pedidos para generar informes de retención y comportamiento de compra.")

# Renderizar el contenido según el modo seleccionado
if mode == "Generar informe":
    st.session_state["show_controls"] = True # Forzar la visualización del panel de carga en este modo

    # ---- Panel de carga y generación ----
    # Ya no es ocultable por el botón, pero se mantiene la estructura para la carga de CSV
    with st.container():
        st.markdown("### Subir datos y generar informes")
        # Asegurarse de que uploaded_file se obtiene de un solo widget
        uploaded_file = st.file_uploader("Elige un archivo CSV", type="csv", key="uploader_panel_main") 

        if uploaded_file is not None and uploaded_file.size > 0: # Añadir verificación de tamaño
            st.session_state["df_raw"] = pd.read_csv(uploaded_file) # Guardar en session_state para persistencia
            preview_placeholder.info("Archivo subido correctamente. Mostrando primeras 5 filas:")
            preview_placeholder.dataframe(st.session_state["df_raw"].head())

            # Selección de formato de fecha (para evitar interpretaciones ambiguas)
            fmt_labels = list(DATE_FORMAT_OPTIONS.keys())
            default_idx = 0
            selected_fmt = st.selectbox(
                "Formato de fecha",
                fmt_labels,
                index=default_idx,
                key="date_format_select"
            )
            st.session_state["date_format_mode"] = DATE_FORMAT_OPTIONS.get(selected_fmt, "auto")

            amount_labels = {
                "Auto (detectar)": "auto",
                "Euros (€)": "eur",
                "Céntimos": "cents",
            }
            selected_amount = st.selectbox(
                "Unidad de Total_pagado",
                list(amount_labels.keys()),
                index=0,
                key="amount_unit_select"
            )
            st.session_state["amount_unit_mode"] = amount_labels.get(selected_amount, "auto")

            fechas_preview = None
            if 'fecha' in st.session_state["df_raw"].columns:
                fechas_preview = parse_date_series(
                    st.session_state["df_raw"]['fecha'],
                    mode=st.session_state["date_format_mode"]
                )
            if (fechas_preview is None or fechas_preview.isna().all()) and 'fecha_hora' in st.session_state["df_raw"].columns:
                fechas_preview = parse_date_series(st.session_state["df_raw"]['fecha_hora'], mode="auto")
                st.info("Usando 'fecha_hora' para detectar el rango porque 'fecha' no es válida.")

            if fechas_preview is None or fechas_preview.isna().all():
                min_fecha = datetime.now() - timedelta(days=365)
                max_fecha = datetime.now()
            else:
                min_fecha = fechas_preview.min()
                max_fecha = fechas_preview.max()
            if pd.isna(min_fecha) or pd.isna(max_fecha):
                min_fecha = datetime.now() - timedelta(days=365)
                max_fecha = datetime.now()
            st.markdown("#### Rango de fechas para procesar")
            range_mode = st.radio(
                "Modo de rango",
                ["Calendario", "Por años"],
                horizontal=True,
                key="process_date_range_mode"
            )
            if range_mode == "Por años":
                years = list(range(min_fecha.year, max_fecha.year + 1))
                if not years:
                    years = [datetime.now().year]
                col_y1, col_y2 = st.columns(2)
                start_year = col_y1.selectbox("Año inicio", years, index=0, key="process_start_year")
                end_year = col_y2.selectbox("Año fin", years, index=len(years) - 1, key="process_end_year")
                if start_year > end_year:
                    st.warning("El año de inicio no puede ser mayor que el año fin. Se ajustó automáticamente.")
                    start_year, end_year = end_year, start_year
                date_start = datetime(start_year, 1, 1).date()
                date_end = datetime(end_year, 12, 31).date()
                st.caption(f"Se analizará desde {date_start} hasta {date_end}.")
            else:
                date_range = st.date_input(
                    "Rango (desde / hasta)",
                    value=(min_fecha.date(), max_fecha.date()),
                    min_value=min_fecha.date(),
                    max_value=max_fecha.date(),
                    key="process_date_range"
                )
                if isinstance(date_range, (list, tuple)) and len(date_range) == 2:
                    date_start, date_end = date_range
                else:
                    date_start = min_fecha.date()
                    date_end = max_fecha.date()

            if st.button("Generar / Actualizar Informes", key="generate_button_main"):
                log_box = status_placeholder.empty()
                log = log_box.write
                with st.spinner("Procesando datos y generando informes... Esto puede tardar unos minutos."):
                    date_mode = st.session_state.get("date_format_mode", "auto")
                    amount_mode = st.session_state.get("amount_unit_mode", "auto")
                    df_processed = preprocess_data(
                        st.session_state["df_raw"].copy(),
                        logger=log,
                        date_mode=date_mode,
                        amount_mode=amount_mode
                    ) # Usar df_raw del session_state
                    if df_processed is not None and not df_processed.empty:
                        start_ts = pd.Timestamp(date_start)
                        end_ts = pd.Timestamp(date_end) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
                        df_processed = df_processed[
                            (df_processed['fecha_dt'] >= start_ts) & (df_processed['fecha_dt'] <= end_ts)
                        ]
                    
                    if df_processed is not None and not df_processed.empty:
                        report1_df = generate_retention_report(df_processed)
                        report2_df = generate_annual_retention_report(df_processed)
                        report3_df, report3_summary, report3_active_window = generate_survival_analysis(df_processed)
                        report4_dfs = generate_frequency_report(df_processed)
                        report5_dfs = generate_loyalty_report(df_processed)

                        all_reports = {
                            "report1": report1_df,
                            "report2": report2_df,
                            "report3": report3_df,
                            "report4": report4_dfs,
                            "report5": report5_dfs,
                        }
                        excel_file = export_to_excel(all_reports) # Se devuelve el io.BytesIO
                        excel_bytes = excel_file.getvalue()

                        # Guardar como base para vista y rango por defecto para la sesión
                        # Si no se guarda persistentemente, sigue siendo visible en la sesión
                        st.session_state["base_reports"] = all_reports
                        st.session_state["base_summary"] = report3_summary
                        st.session_state["base_active_window"] = report3_active_window
                        st.session_state["data_date_min"] = df_processed['fecha_dt'].min().date()
                        st.session_state["data_date_max"] = df_processed['fecha_dt'].max().date()
                        st.session_state["view_date_range"] = (st.session_state["data_date_min"], st.session_state["data_date_max"])

                        # Persistir último informe generado para poder guardarlo en un rerun
                        st.session_state["last_generated_excel_bytes"] = excel_bytes
                        st.session_state["last_generated_reports"] = all_reports
                        st.session_state["last_generated_summary"] = report3_summary
                        st.session_state["last_generated_active_window"] = report3_active_window
                        st.session_state["last_generated_range"] = (date_start, date_end)
                        st.session_state["save_report_name_input"] = f"Informe CLV {datetime.now().strftime('%Y%m%d_%H%M')}"

                        # Limpiar vista previa y logs;
                        log_box.empty()
                        preview_placeholder.empty()
                        status_placeholder.empty()

                        st.sidebar.download_button(
                            label="📥 Descargar Informes en Excel (sesión actual)",
                            data=excel_bytes,
                            file_name=f"Analisis_CLV_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_latest_session"
                        )
                    else:
                        st.error("No se pudieron procesar los datos. Revisa el CSV o el rango de fechas.")
        elif uploaded_file is not None and uploaded_file.size == 0:
            st.error("El archivo CSV subido está vacío. Por favor, sube un archivo con datos.")
        elif uploaded_file is None:
            st.info("Sube un archivo CSV para comenzar el análisis.")

    # --- Guardar informe (persistente) fuera del botón de generación ---
    if st.session_state.get("last_generated_excel_bytes"):
        st.subheader("Guardar Informe Generado")
        report_name_input = st.text_input(
            "Introduce un nombre para guardar este informe:",
            value=st.session_state.get("save_report_name_input", f"Informe CLV {datetime.now().strftime('%Y%m%d_%H%M')}"),
            key="save_report_name_input"
        )

        if st.button("💾 Guardar Informe", key="save_report_button"):
            if report_name_input:
                reports_index = load_reports_index()
                if report_name_input in reports_index:
                    st.warning(f"Ya existe un informe con el nombre '{report_name_input}'. Por favor, elige otro nombre.")
                else:
                    report_filepath = get_report_filepath(report_name_input)
                    with open(report_filepath, "wb") as f:
                        f.write(st.session_state["last_generated_excel_bytes"])

                    date_start, date_end = st.session_state.get("last_generated_range", (None, None))
                    json_safe_summary = make_json_safe(st.session_state.get("last_generated_summary", {}))
                    active_window = st.session_state.get("last_generated_active_window") or {}
                    active_start = active_window.get("start")
                    active_end = active_window.get("end")
                    reports_index[report_name_input] = {
                        "filename": os.path.basename(report_filepath),
                        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M"),
                        "date_min": date_start.isoformat() if date_start else "",
                        "date_max": date_end.isoformat() if date_end else "",
                        "summary": json_safe_summary,
                        "active_window_start": active_start.isoformat() if active_start else "",
                        "active_window_end": active_end.isoformat() if active_end else ""
                    }
                    save_reports_index(reports_index)
                    st.success(f"Informe '{report_name_input}' guardado exitosamente.")
                    st.session_state["selected_report"] = report_name_input
                    safe_rerun()
            else:
                st.warning("Por favor, introduce un nombre para el informe.")

        if st.session_state.get("base_reports"):
            if st.button("Ver informe ahora", key="go_view_report"):
                st.session_state["nav_to"] = "Ver informe"
                safe_rerun()




elif mode == "Informes guardados":
    st.session_state["show_controls"] = False # Ocultar el panel de carga en este modo
    st.markdown("### Informes CLV Guardados")
    
    reports_index = load_reports_index()
    if not reports_index:
        st.info("No hay informes guardados aún. Genera uno nuevo para empezar.")
    else:
        # Ordenar informes por fecha de guardado (más reciente primero)
        sorted_reports_items = sorted(reports_index.items(), key=lambda item: item[1]['timestamp'], reverse=True)
        
        if st.session_state.get("selected_report"):
            st.caption(f"Informe cargado actualmente: {st.session_state['selected_report']}")

        header = st.columns([3, 2, 3, 1.3, 1.7, 1.2])
        header[0].markdown("**Informe**")
        header[1].markdown("**Guardado el**")
        header[2].markdown("**Rango de datos**")
        header[3].markdown("**Cargar**")
        header[4].markdown("**Descargar**")
        header[5].markdown("**Eliminar**")

        for i, (report_name, report_data) in enumerate(sorted_reports_items):
            row = st.columns([3, 2, 3, 1.3, 1.7, 1.2])
            row[0].write(report_name)
            row[1].write(report_data.get("timestamp", ""))
            row[2].write(f"{report_data.get('date_min', '')} a {report_data.get('date_max', '')}")

            with row[3]:
                if st.button("Cargar", key=f"load_report_{i}"):
                    try:
                        report_filepath = os.path.join(REPORTS_DIR, report_data["filename"])
                        if os.path.exists(report_filepath):
                            all_reports_from_excel = pd.read_excel(report_filepath, sheet_name=None, index_col=0)
                            loaded_reports = {
                                "report1": all_reports_from_excel.get('Retención Trimestral'),
                                "report2": all_reports_from_excel.get('Retención Anual'),
                                "report3": all_reports_from_excel.get('Análisis de Supervivencia'),
                                "report4": {
                                    "distribucion": all_reports_from_excel.get('Frecuencia - Distribución'),
                                    "segunda_compra": all_reports_from_excel.get('Frecuencia - 2da Compra'),
                                    "evolucion": all_reports_from_excel.get('Frecuencia - Evolución'),
                                    "velocidad": all_reports_from_excel.get('Frecuencia - Velocidad'),
                                }
                                ,
                                "report5": {
                                    "activacion": all_reports_from_excel.get('Lealtad - Activación'),
                                    "repeticion": all_reports_from_excel.get('Lealtad - Repetición'),
                                    "ingresos_recurrentes": all_reports_from_excel.get('Lealtad - Ingresos Recurrentes'),
                                    "repeticion_pedidos": all_reports_from_excel.get('Lealtad - Repetición vs Pedidos'),
                                }
                            }
                            if loaded_reports.get("report3") is not None and 'Total Clientes' in loaded_reports["report3"].columns:
                                loaded_reports["report3"]['Total Clientes'] = loaded_reports["report3"]['Total Clientes'].fillna(0).astype(int)

                            st.session_state["base_reports"] = loaded_reports
                            st.session_state["base_summary"] = report_data.get("summary", {})
                            aw_start = report_data.get("active_window_start")
                            aw_end = report_data.get("active_window_end")
                            if aw_start and aw_end:
                                st.session_state["base_active_window"] = {
                                    "start": datetime.fromisoformat(aw_start).date(),
                                    "end": datetime.fromisoformat(aw_end).date(),
                                }
                            else:
                                st.session_state["base_active_window"] = None
                            st.session_state["data_date_min"] = datetime.fromisoformat(report_data['date_min']).date() if report_data.get('date_min') else None
                            st.session_state["data_date_max"] = datetime.fromisoformat(report_data['date_max']).date() if report_data.get('date_max') else None
                            if st.session_state["data_date_min"] and st.session_state["data_date_max"]:
                                st.session_state["view_date_range"] = (st.session_state["data_date_min"], st.session_state["data_date_max"])
                            st.session_state["selected_report"] = report_name
                            st.success(f"Informe '{report_name}' cargado para visualización.")
                            st.session_state["nav_to"] = "Ver informe"
                            safe_rerun()
                        else:
                            st.error(f"Archivo de informe '{report_data['filename']}' no encontrado.")
                    except Exception as e:
                        st.error(f"Error al cargar el informe '{report_name}': {e}")

            with row[4]:
                report_filepath = os.path.join(REPORTS_DIR, report_data["filename"])
                if os.path.exists(report_filepath):
                    with open(report_filepath, "rb") as f:
                        download_data = f.read()
                    st.download_button(
                        label="Descargar",
                        data=download_data,
                        file_name=report_data["filename"],
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"download_saved_excel_{i}"
                    )
                else:
                    st.caption("No encontrado")

            with row[5]:
                if st.button("Eliminar", key=f"delete_saved_{i}"):
                    st.session_state["delete_candidate"] = report_name

            if st.session_state.get("delete_candidate") == report_name:
                st.warning(f"¿Confirmas eliminar el informe '{report_name}' y su archivo asociado?")
                del_col1, del_col2 = st.columns(2)
                with del_col1:
                    if st.button("Confirmar eliminación", key=f"confirm_delete_{i}"):
                        try:
                            report_filepath = os.path.join(REPORTS_DIR, report_data["filename"])
                            if os.path.exists(report_filepath):
                                os.remove(report_filepath)

                            del reports_index[report_name]
                            save_reports_index(reports_index)
                            st.success(f"Informe '{report_name}' eliminado exitosamente.")
                            if st.session_state.get("selected_report") == report_name:
                                st.session_state["selected_report"] = None
                                st.session_state["base_reports"] = None
                            st.session_state["delete_candidate"] = None
                            safe_rerun()
                        except Exception as e:
                            st.error(f"Error al eliminar el informe '{report_name}': {e}")
                with del_col2:
                    if st.button("Cancelar", key=f"cancel_delete_{i}"):
                        st.session_state["delete_candidate"] = None


# --- Comparación de informes guardados ---
elif mode == "Comparar informes":
    reports_index = load_reports_index()
    if not reports_index or len(reports_index) < 2:
        st.info("Necesitas al menos 2 informes guardados para comparar.")
    else:
        st.subheader("Comparar informes guardados")
        sorted_reports_items = sorted(reports_index.items(), key=lambda item: item[1]['timestamp'], reverse=True)
        report_names = [name for name, _ in sorted_reports_items]

        default_a = 0
        default_b = 1 if len(report_names) > 1 else 0
        report_a = st.selectbox("Informe A", report_names, index=default_a, key="compare_report_a")
        report_b = st.selectbox("Informe B", report_names, index=default_b, key="compare_report_b")

        if report_a == report_b:
            st.warning("Selecciona dos informes distintos para comparar.")
        else:
            data_a = reports_index.get(report_a, {})
            data_b = reports_index.get(report_b, {})

            st.markdown(
                f"**A:** {report_a} | Guardado: {data_a.get('timestamp', '')} | Rango: {data_a.get('date_min', '')} a {data_a.get('date_max', '')}"
            )
            st.markdown(
                f"**B:** {report_b} | Guardado: {data_b.get('timestamp', '')} | Rango: {data_b.get('date_min', '')} a {data_b.get('date_max', '')}"
            )
            st.caption("Diferencias calculadas como: B − A (puntos porcentuales en retención).")

            def load_saved_report(report_data):
                report_filepath = os.path.join(REPORTS_DIR, report_data.get("filename", ""))
                if not os.path.exists(report_filepath):
                    return None
                all_reports = pd.read_excel(report_filepath, sheet_name=None, index_col=0)
                return {
                    "report1": all_reports.get('Retención Trimestral'),
                    "report2": all_reports.get('Retención Anual'),
                }

            reports_a = load_saved_report(data_a)
            reports_b = load_saved_report(data_b)

            if reports_a is None:
                st.error(f"No se pudo leer el archivo del informe A: {data_a.get('filename', '')}")
            if reports_b is None:
                st.error(f"No se pudo leer el archivo del informe B: {data_b.get('filename', '')}")

            if reports_a is not None and reports_b is not None:
                # --- Retención anual ---
                annual_a, annual_b = align_retention_tables(reports_a.get("report2"), reports_b.get("report2"), freq="Y")
                annual_delta = None
                if annual_a is not None and annual_b is not None:
                    annual_delta = annual_b - annual_a

                st.markdown("### Retención anual por cohortes")
                if annual_a is None or annual_b is None:
                    st.warning("No se encontró el informe anual en uno de los archivos seleccionados.")
                else:
                    tabs_annual = st.tabs(["Informe A", "Informe B", "Diferencias"])
                    with tabs_annual[0]:
                        show_table(annual_a, style_percent_heatmap, info_msg="Vista sin estilo por compatibilidad.")
                    with tabs_annual[1]:
                        show_table(annual_b, style_percent_heatmap, info_msg="Vista sin estilo por compatibilidad.")
                    with tabs_annual[2]:
                        show_table(annual_delta, style_delta_retention, info_msg="Vista sin estilo por compatibilidad.")

                    with st.expander("Detalle anual (conteos estimados por cohorte)"):
                        counts_a = retention_counts_from_pct(annual_a)
                        counts_b = retention_counts_from_pct(annual_b)
                        counts_delta = counts_b - counts_a
                        tabs_counts = st.tabs(["Informe A", "Informe B", "Diferencias"])
                        with tabs_counts[0]:
                            show_table(counts_a, lambda d: style_heatmap(d, cmap="Blues"), info_msg="Vista sin estilo por compatibilidad.")
                        with tabs_counts[1]:
                            show_table(counts_b, lambda d: style_heatmap(d, cmap="Blues"), info_msg="Vista sin estilo por compatibilidad.")
                        with tabs_counts[2]:
                            show_table(counts_delta, style_delta_counts, info_msg="Vista sin estilo por compatibilidad.")

                # --- Retención trimestral ---
                quarterly_a, quarterly_b = align_retention_tables(reports_a.get("report1"), reports_b.get("report1"), freq="Q")
                quarterly_delta = None
                if quarterly_a is not None and quarterly_b is not None:
                    quarterly_delta = quarterly_b - quarterly_a

                st.markdown("### Retención trimestral por cohortes")
                if quarterly_a is None or quarterly_b is None:
                    st.warning("No se encontró el informe trimestral en uno de los archivos seleccionados.")
                else:
                    tabs_quarter = st.tabs(["Informe A", "Informe B", "Diferencias"])
                    with tabs_quarter[0]:
                        show_table(quarterly_a, style_retention_table, info_msg="Vista sin estilo por compatibilidad.")
                    with tabs_quarter[1]:
                        show_table(quarterly_b, style_retention_table, info_msg="Vista sin estilo por compatibilidad.")
                    with tabs_quarter[2]:
                        show_table(quarterly_delta, style_delta_retention, info_msg="Vista sin estilo por compatibilidad.")

# --- Visualización de los informes (solo en modo "Ver informe") ---
if mode == "Ver informe":
    base_reports = st.session_state.get("base_reports")
    base_summary = st.session_state.get("base_summary")
    base_active_window = st.session_state.get("base_active_window")

    if base_reports:
        st.markdown(f"## Informe: {st.session_state.get('selected_report', 'Recién Generado')}")
        data_min = st.session_state.get("data_date_min")
        data_max = st.session_state.get("data_date_max")
        if not data_min or not data_max or (data_min and data_max and data_min > data_max):
            today = datetime.now().date()
            data_min = today - timedelta(days=365)
            data_max = today
        view_range_value = st.session_state.get("view_date_range", (data_min, data_max))
        if not isinstance(view_range_value, (list, tuple)) or len(view_range_value) != 2:
            view_range_value = (data_min, data_max)
        view_start, view_end = view_range_value
        if isinstance(view_start, datetime):
            view_start = view_start.date()
        if isinstance(view_end, datetime):
            view_end = view_end.date()
        if not view_start or not view_end:
            view_start, view_end = data_min, data_max
        if view_start < data_min:
            view_start = data_min
        if view_end > data_max:
            view_end = data_max
        if view_start > view_end:
            view_start, view_end = data_min, data_max
        st.session_state["view_date_range"] = (view_start, view_end)

        with st.sidebar.expander("Filtrar visualización por fechas", expanded=False):
            view_range_mode = st.radio(
                "Modo de rango",
                ["Calendario", "Por años"],
                horizontal=True,
                key="view_date_range_mode"
            )
            if view_range_mode == "Por años":
                years = list(range(data_min.year, data_max.year + 1))
                if not years:
                    years = [datetime.now().year]
                col_y1, col_y2 = st.columns(2)
                start_year = col_y1.selectbox("Año inicio", years, index=0, key="view_start_year")
                end_year = col_y2.selectbox("Año fin", years, index=len(years) - 1, key="view_end_year")
                if start_year > end_year:
                    st.warning("El año de inicio no puede ser mayor que el año fin. Se ajustó automáticamente.")
                    start_year, end_year = end_year, start_year
                view_range = (datetime(start_year, 1, 1).date(), datetime(end_year, 12, 31).date())
                st.caption(f"Mostrando desde {view_range[0]} hasta {view_range[1]}.")
            else:
                view_range = st.date_input(
                    "Rango de visualización",
                    value=st.session_state.get("view_date_range", (data_min, data_max)),
                    min_value=data_min,
                    max_value=data_max,
                    key="view_date_range"
                )
        if isinstance(view_range, (list, tuple)) and len(view_range) == 2:
            view_start, view_end = view_range
        else:
            view_start, view_end = data_min, data_max

        display_reports = filter_reports_by_date(base_reports, view_start, view_end)
        tab1, tab2, tab3, tab4 = st.tabs([
            "Informe 1: Retención Trimestral",
            "Informe 2: Retención Anual",
            "Informe 3: Análisis de Supervivencia",
            "Informe 4: Frecuencia de Compra"
        ])

        with tab1:
            st.header("Retención por Trimestres")
            def _help_ret_trimestral():
                st.markdown(
                    "Análisis de cohortes que muestra qué porcentaje de clientes de cada cohorte trimestral "
                    "(basada en su primera compra) realiza compras en trimestres subsiguientes."
                )
                st.markdown(
                    "- **Filas**: Cohorte (trimestre de primera compra)\n"
                    "- **Columnas**: Trimestres calendario\n"
                    "- **Valores**: % de retención"
                )
                st.markdown(
                    "Cada celda indica el **porcentaje de clientes** de la cohorte (fila) que realizaron **al menos una compra** "
                    "en el periodo (columna). No es porcentaje de pedidos."
                )
                st.markdown(
                    "Ejemplo: si en la fila `Y2020-Q4` y la columna `Y2023-Q3` aparece `2.4%`, "
                    "significa que **el 2.4% de los clientes cuya primera compra fue en Y2020‑Q4 compraron al menos una vez en Y2023‑Q3**."
                )
                st.markdown(
                    "Pistas rápidas: la diagonal suele ser `100%` (primera compra) y los periodos **anteriores a la cohorte** aparecen como `0%`."
                )
            help_popup("❓ Ayuda", _help_ret_trimestral)
            if display_reports.get('report1') is not None and not display_reports['report1'].empty:
                show_table(
                    display_reports['report1'],
                    styler_fn=lambda d: style_percent_heatmap(d, cmap=RETENTION_CMAP),
                    info_msg="Tu versión de Streamlit no soporta estilos de pandas (<1.31). Se muestra la tabla sin colores."
                )

        with tab2:
            st.header("Retención Anual")
            def _help_ret_anual():
                st.markdown("Versión agregada del análisis de retención a nivel anual.")
                st.markdown(
                    "Cada celda indica el **porcentaje de clientes** de la cohorte (fila) que realizaron **al menos una compra** "
                    "en el año (columna). No es porcentaje de pedidos."
                )
                st.markdown(
                    "Ejemplo: si en la fila `2020` y la columna `2023` aparece `2.4%`, "
                    "significa que **el 2.4% de los clientes cuya primera compra fue en 2020 compraron al menos una vez en 2023**."
                )
                st.markdown(
                    "Pistas rápidas: la diagonal suele ser `100%` y los años **anteriores a la cohorte** aparecen como `0%`."
                )
            help_popup("❓ Ayuda", _help_ret_anual)
            if display_reports.get('report2') is not None and not display_reports['report2'].empty:
                 show_table(
                     display_reports['report2'],
                     styler_fn=lambda d: style_percent_heatmap(d, cmap=RETENTION_CMAP),
                     info_msg="Tu versión de Streamlit no soporta estilos de pandas (<1.31). Se muestra la tabla sin colores."
                 )

        with tab3:
            st.header("Análisis de Supervivencia")
            def _help_survival():
                st.markdown(
                    "**Qué mide**: la tabla muestra el porcentaje de clientes de cada cohorte que siguen "
                    "\"vivos\" (han comprado al menos una vez) **a partir de** un número de meses desde su primera compra."
                )
                st.markdown(
                    "**Cómo leerla**:\n"
                    "- **Fila** = cohorte (trimestre de primera compra).\n"
                    "- **Mes X** = % de clientes que **realizaron al menos una compra en o después** de X meses.\n"
                    "- **Total Clientes** = tamaño de la cohorte."
                )
                st.markdown(
                    "**Ejemplo 1**: si en `2020‑Q4` el valor en **Mes 12** es `22.76%`, significa que "
                    "**el 22.76% de los clientes que compraron por primera vez en 2020‑Q4 hicieron "
                    "alguna compra a partir de los 12 meses** desde su primera compra."
                )
                st.markdown(
                    "**Ejemplo 2**: si en `2021‑Q2` el valor en **Mes 36** es `3.49%`, significa que "
                    "**solo el 3.49% de esa cohorte sigue comprando al menos una vez después de 36 meses**."
                )
                st.markdown(
                    "**Lectura rápida**:\n"
                    "- Los valores deben **disminuir** a medida que aumenta el mes.\n"
                    "- Un **0%** indica que no hay clientes de esa cohorte con compras después de ese umbral."
                )
                st.markdown(
                    "**Columnas finales**:\n"
                    "- **Lifetime_Prom**: días promedio entre primera y última compra.\n"
                    "- **Pedidos_Prom**: pedidos promedio por cliente.\n"
                    "- **Revenue_Prom**: ingresos promedio por cliente."
                )
            help_popup("❓ Ayuda", _help_survival)
            if display_reports.get('report3') is not None and not display_reports['report3'].empty:
                st.subheader("Resumen Ejecutivo")
                if base_summary:
                    cols = st.columns(len(base_summary))
                    for i, (key, value) in enumerate(base_summary.items()):
                        if "%" in key:
                            cols[i].metric(key, f"{value:.2f}%")
                        elif key.startswith("Clientes activos"):
                            cols[i].metric(key, f"{int(value):,}")
                        elif "días" in key or "Promedio" in key:
                            cols[i].metric(key, f"{value:.2f}")
                        else:
                            cols[i].metric(key, f"{int(value):,}")
                if base_active_window:
                    st.caption(
                        f"Ventana usada para 'Clientes activos (últimos 90 días)': "
                        f"{base_active_window['start']} a {base_active_window['end']}"
                    )

                st.subheader("Tabla de Supervivencia por Cohorte")
                show_table(
                    display_reports['report3'],
                    styler_fn=style_survival_table,
                    info_msg="Tu versión de Streamlit no soporta estilos de pandas (<1.31). Se muestra la tabla sin colores."
                )

        with tab4:
            st.header("Frecuencia de Compra")
            def _help_freq():
                st.markdown(
                    "**Qué analiza**: comportamiento de recompra de clientes con **2 o más pedidos**. "
                    "Se centra en el tiempo entre compras y en la velocidad de compra."
                )
                st.markdown(
                    "**Notas clave**:\n"
                    "- Las métricas de intervalos **no cuentan pedidos**, sino **intervalos entre compras**.\n"
                    "- En \"2ª compra\" solo entran clientes que **sí tienen** segunda compra.\n"
                    "- En \"Evolución\" el intervalo de la compra N es el tiempo entre la compra N‑1 y N."
                )
                st.markdown(
                    "**Ejemplo**: si un cliente compra el 1/01, 15/01 y 01/02, entonces:\n"
                    "- Intervalos: 14 días y 17 días.\n"
                    "- En \"Evolución\", el intervalo de compra 2 es 14 días y el de compra 3 es 17 días."
                )
                st.markdown(
                    "**Velocidad de compra**: se calcula como `total_pedidos / (días_actividad/30)`. "
                    "Luego se agrupa en segmentos (Muy baja, Baja, Media, etc.)."
                )
                st.markdown(
                    "**Frecuencia y Recurrencia (Lealtad)**:\n"
                    "- **Tasa de activación** = `COUNT_DISTINCT(IF(Num_Pedidos_Cliente = 2, cod_cliente)) / "
                    "COUNT_DISTINCT(IF(Num_Pedidos_Cliente = 1, cod_cliente))`\n"
                    "- **Tasa de repetición** = `COUNT_DISTINCT(CASE WHEN cliente_nuevo = \"n\" THEN email END) / "
                    "COUNT_DISTINCT(email)`\n"
                    "- **Tasa ingresos recurrentes** = `SUM(IF(cliente_nuevo = \"n\", Total_pagado, 0)) / SUM(Total_pagado)`\n"
                    "- **Total pedidos** = `COUNT_DISTINCT(codigo)`\n\n"
                    "Se calcula **por año de la fecha del pedido** y respeta el rango de fechas procesado."
                )
            help_popup("❓ Ayuda", _help_freq)

            if display_reports and display_reports.get('report5'):
                st.subheader("Frecuencia y Recurrencia (Lealtad)")
                r5 = display_reports['report5']

                col1, col2 = st.columns(2)
                with col1:
                    st.markdown("**Tasa de activación**")
                    df_act = r5.get('activacion')
                    if df_act is not None and not df_act.empty:
                        fig = plot_bar_with_labels(df_act, "Tasa_activacion", color="#4C8BF5", is_percent=True)
                        if fig is not None:
                            st.pyplot(fig, clear_figure=True)
                        else:
                            st.bar_chart(df_act)
                    show_table(
                        df_act,
                        styler_fn=style_percent_heatmap,
                        info_msg="Tu versión de Streamlit no soporta estilos de pandas (<1.31). Se muestra la tabla sin colores."
                    )
                with col2:
                    st.markdown("**Tasa de repetición**")
                    df_rep = r5.get('repeticion')
                    if df_rep is not None and not df_rep.empty:
                        fig = plot_bar_with_labels(df_rep, "Tasa_repeticion", color="#4C8BF5", is_percent=True)
                        if fig is not None:
                            st.pyplot(fig, clear_figure=True)
                        else:
                            st.bar_chart(df_rep)
                    show_table(
                        df_rep,
                        styler_fn=style_percent_heatmap,
                        info_msg="Tu versión de Streamlit no soporta estilos de pandas (<1.31). Se muestra la tabla sin colores."
                    )

                col3, col4 = st.columns(2)
                with col3:
                    st.markdown("**Tasa de ingresos de clientes recurrentes**")
                    df_rev = r5.get('ingresos_recurrentes')
                    if df_rev is not None and not df_rev.empty:
                        fig = plot_bar_with_labels(df_rev, "Tasa_ingresos_recurrentes", color="#4C8BF5", is_percent=True)
                        if fig is not None:
                            st.pyplot(fig, clear_figure=True)
                        else:
                            st.bar_chart(df_rev)
                    show_table(
                        df_rev,
                        styler_fn=style_percent_heatmap,
                        info_msg="Tu versión de Streamlit no soporta estilos de pandas (<1.31). Se muestra la tabla sin colores."
                    )
                with col4:
                    st.markdown("**Tasa de repetición y total pedidos**")
                    df_rep_ped = r5.get('repeticion_pedidos')
                    if df_rep_ped is not None and not df_rep_ped.empty:
                        fig = plot_repetition_vs_orders(df_rep_ped)
                        if fig is not None:
                            st.pyplot(fig, clear_figure=True)
                        else:
                            st.bar_chart(df_rep_ped)
                    show_table(
                        df_rep_ped,
                        styler_fn=style_repetition_orders,
                        info_msg="Tu versión de Streamlit no soporta estilos de pandas (<1.31). Se muestra la tabla sin colores."
                    )

            if display_reports and display_reports.get('report4'):
                ft1, ft2, ft3, ft4 = st.tabs([
                    "Distribución",
                    "2ª compra",
                    "Evolución",
                    "Velocidad"
                ])

                with ft1:
                    st.subheader("Distribución por Frecuencia de Compra")
                    st.caption("Intervalos de días entre compras para clientes con 2+ pedidos.")
                    df_dist = display_reports['report4']['distribucion']
                    if df_dist is not None and not df_dist.empty:
                        chart_df = df_dist[['% del Total']].copy() if '% del Total' in df_dist.columns else df_dist.copy()
                        st.bar_chart(chart_df)
                    show_table(
                        df_dist,
                        styler_fn=style_frequency_distribucion,
                        info_msg="Tu versión de Streamlit no soporta estilos de pandas (<1.31). Se muestra la tabla sin colores."
                    )

                with ft2:
                    st.subheader("Tiempo hasta la Segunda Compra")
                    st.caption("Clientes que vuelven a comprar según el tiempo transcurrido desde su primera compra.")
                    df_2 = display_reports['report4']['segunda_compra']
                    if df_2 is not None and not df_2.empty:
                        chart_df = df_2[['% del Total']].copy() if '% del Total' in df_2.columns else df_2.copy()
                        st.bar_chart(chart_df)
                    show_table(
                        df_2,
                        styler_fn=style_frequency_second_purchase,
                        info_msg="Tu versión de Streamlit no soporta estilos de pandas (<1.31). Se muestra la tabla sin colores."
                    )

                with ft3:
                    st.subheader("Evolución de Frecuencia por Número de Compra")
                    st.caption("Cómo cambia el intervalo entre compras a medida que aumenta el número de pedido.")
                    df_evo = display_reports['report4']['evolucion']
                    if df_evo is not None and not df_evo.empty:
                        line_cols = [c for c in ['Dias_Promedio_Intervalo', 'Dias_Mediana_Intervalo'] if c in df_evo.columns]
                        if line_cols:
                            st.line_chart(df_evo[line_cols])
                    show_table(
                        df_evo,
                        styler_fn=style_frequency_evolucion,
                        info_msg="Tu versión de Streamlit no soporta estilos de pandas (<1.31). Se muestra la tabla sin colores."
                    )

                with ft4:
                    st.subheader("Velocidad de Compra (Compras por Mes)")
                    st.caption("Segmentación por compras/mes con métricas promedio por cliente.")
                    df_vel = display_reports['report4']['velocidad']
                    if df_vel is not None and not df_vel.empty:
                        chart_df = df_vel[['% del Total']].copy() if '% del Total' in df_vel.columns else df_vel.copy()
                        st.bar_chart(chart_df)
                    show_table(
                        df_vel,
                        styler_fn=style_frequency_velocidad,
                        info_msg="Tu versión de Streamlit no soporta estilos de pandas (<1.31). Se muestra la tabla sin colores."
                    )
    else:
        st.info("Carga un informe guardado o genera uno nuevo para visualizarlo.")
